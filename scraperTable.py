from pprint import pp, pprint
from openpyxl import Workbook
from openpyxl.worksheet.page import PrintOptions

# Setup
wb = Workbook()
ws = wb.active
ws.title = 'Get Scraping'
filename = 'example-scraper.xlsx'

def insert_scraped_data(c, data):
  c.value = data

def create_titles(): 
  ws['A1'] = 'Name'
  ws['B1'] = 'Email'
  ws['C1'] = 'Website'
  ws['D1'] = 'Location'
  ws['E1'] = 'About'

def set_column_dimensions(): 
  columns = ['A', 'B', 'C', 'D', 'E']
  for col in columns:
    ws.column_dimensions[col].width = 30  
  return

def populate_sheet(data):
  rows = ws.iter_rows(min_row=2, max_row=len(data), max_col=len(data[0]))
  create_titles()
  set_column_dimensions()
  i = 0
  for  row in rows:
    i = i+1
    for cell in row:
      if cell.column == 1: insert_scraped_data(cell, data[i - 1]['name'])
      if cell.column == 2: insert_scraped_data(cell, data[i - 1]['email'])
      if cell.column == 3: insert_scraped_data(cell, data[i - 1]['website'])
      if cell.column == 4: insert_scraped_data(cell, data[i - 1]['address'])
      if cell.column == 5: 
        insert_scraped_data(cell, data[i - 1]['about'])
    
    wb.save(filename=filename)
  return
