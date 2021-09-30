from openpyxl import Workbook
from pprint import pp, pprint

filename = "practice_sheet.xlsx"
wb = Workbook()
ws = wb.active

def print_rows():
  for row in ws.iter_rows(values_only=True):
    pprint(row)

ws['A1'] = 'Hello'
ws['B1'] = 'World'


cell = ws['A1']
# cell.value = 'Bitch'


ws.insert_cols(idx=1)
# (None, 'hello', 'world!')

ws.insert_cols(idx=3, amount=5)
# (None, 'hello', None, None, None, None, None, 'world!')

ws.delete_cols(idx=3, amount=5)
ws.delete_cols(idx=1)
# ('hello', 'world!')

ws.insert_rows(idx=1)
# (None, None)
# ('hello', 'world!')

ws.insert_rows(idx=1, amount=3)
# (None, None)
# (None, None)
# (None, None)
# (None, None)
# ('hello', 'world!')

ws.delete_rows(idx=1, amount=4)

ws.insert_rows(idx=1, amount=5)

wb.save(filename=filename)