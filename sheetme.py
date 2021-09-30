from pprint import pprint
from openpyxl import Workbook
from openpyxl import worksheet
import datetime

wb = Workbook()
ws = wb.active
# Assigns the Title of Hello World to the active worksheet
ws.title = 'Hello World'
# Looks for worksheet with the name Hello World
ws1 = wb['Hello World']
# Create Sheet - Appends to the List by Default
ws2 = wb.create_sheet('Python is Cool')

# Modifying a Cell - Access by c.value
c = ws['A4']
ws['A4'] = 4

# Accesing a range of cells
cell_range = ws2['A1':'D2']
# Access Collum
column_c = ws['A']
column_range = ws['C':'D']
row_ten = ws[10]
row_range = ws[5:10]
# pprint(row_range)

# Iterate through the Rows
def iterate_rows():
  for row in ws.iter_rows(min_row=1, max_col=3, max_row=4):
    for cell in row:
      pprint(cell)
# for every 'row' definded by iter_rows(minimum row= row-1, how many columns= 10, where do you want )
# Prints each cell in a row - A1, B1, C1 => A2, B2, C2

# Iterate through the Columns
def iterate_columns():
  for col in ws.iter_cols(min_row=1, max_col=3, max_row=4):
    for cell in col:
      print(cell)
# Prints each cell in a column A1, A2, A3 - B1, B2, B3, C1 C2 C3

def Print_All_Cells():
  pprint(tuple(ws.rows))

def Print_All_Columns():
  pprint(tuple(ws.rows))

def Print_Value():
  pprint(c.value)





