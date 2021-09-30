import smtplib, ssl
from openpyxl import Workbook, workbook
from openpyxl import load_workbook

wb = load_workbook(filename="testfile.xlsx")
ws = wb.active
array = []
i = 0
for row in ws.iter_rows(min_row=2, max_row=4, max_col=3):
  obj = {} 
  array.append(obj)
  for cell in row:
    if cell.column == 1: obj['Name'] = cell.value
    if cell.column == 2: obj['Email'] = cell.value
    if cell.column == 3: obj['Msg'] = cell.value

array.pop()

# print(array)

port = 465  # For SSL
smtp_server = "smtp.gmail.com"
for i in array:
  name = i['Name']
  msg = i['Msg']
  sender_email = "testbambridgewebscraper@gmail.com"  # Enter your address
  receiver_email = i['Email']  # Enter receiver address
  password = 'Heery123!'
  message = """\
  Subject: Hi there\n\n
  Your name is {name} and this is your message: {msg}"""
  context = ssl.create_default_context()
  with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_email, message.format(name=name, msg=msg))