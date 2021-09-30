from pprint import pp, pprint
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'Send Message to Employees'
filename = 'Employees.xlsx'
testfile = 'testfile.xlsx'

def insert_scraped_data(c, data):
  c.value = data
test = [
    {
      'Subject': 'Daniel',
      'Email': 'danieljheery@gmail.com',
      'Msg': 'Difference are what make us beautiful',
   },
    {
      'Subject': 'Daniel',
      'Email': 'danielheery@bambridgeaccountants.co.uk',
      'Msg': 'You are a beautiful man',
   },
    {
      'Subject': 'Daniel',
      'Email': 'heery@live.co.uk',
      'Msg': 'Eat My poo',
   },
    {
      'Subject': 'Daniel',
      'Email': 'heery@live.co.uk',
      'Msg': 'This is the message',
   },
  ]
Employees = [
  {
    'Name' : 'Anna Heery',
    'Role' : 'Partner',
    'Email': 'anna@bambridgeaccountants.co.uk',
    'Msg' : 'Did you know you are related to Daniel?'
  },
  {
    'Name' : 'Jess',
    'Role' : 'Employee Engagement ',
    'Email': 'jessicawelhenage@bambridgeaccountants.co.uk',
    'Msg' : 'I Think you really love dogs?'
  },
  {
    'Name' : 'Daniel Heery',
    'Role' : 'Web Development & Marketing Associate',
    'Email' : 'danielheery@bambridgeaccountants.co.uk',
    'Msg' : 'Do I know you from somewhere?'
  },
  {
    'Name' : 'Daniel Heery',
    'Role' : 'Web Development & Marketing Associate',
    'Email' : 'danielheery@bambridgeaccountants.co.uk',
    'Msg' : 'Do I know you from somewhere?'
  }
]

def create_titles():
  ws['A1'] = 'Name'
  ws['B1'] = 'Role'
  ws['C1'] = 'Email'
  ws['D1'] = 'Msg'

def set_column_dimensions(): 
  columns = ['A', 'B', 'C', 'D', 'E']
  for col in columns:
    ws.column_dimensions[col].width = 30  
  return

def populate_sheet(data):
  rows = ws.iter_rows(min_row=2, max_row=len(data), max_col=len(data[0]))
  create_titles()
  set_column_dimensions()
  i = 0
  for  row in rows:
    i = i+1
    for cell in row:
      if cell.column == 1: insert_scraped_data(cell, data[i - 1]['Name'])
      if cell.column == 2: insert_scraped_data(cell, data[i - 1]['Role'])
      if cell.column == 3: insert_scraped_data(cell, data[i - 1]['Email'])
      if cell.column == 4: insert_scraped_data(cell, data[i - 1]['Msg'])
      if cell.column == 5: 
        insert_scraped_data(cell, data[i - 1]['about'])

def populate_test_sheet(data):
  rows = ws.iter_rows(min_row=2, max_row=len(data), max_col=len(data[0]))
  create_titles()
  set_column_dimensions()
  i = 0
  for  row in rows:
    i = i+1
    for cell in row:
      if cell.column == 1: insert_scraped_data(cell, data[i - 1]['Subject'])
      if cell.column == 2: insert_scraped_data(cell, data[i - 1]['Email'])
      if cell.column == 3: insert_scraped_data(cell, data[i - 1]['Msg'])

populate_test_sheet(test)
wb.save(filename=testfile)